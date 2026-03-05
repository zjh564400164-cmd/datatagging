from __future__ import annotations

from copy import copy
from io import BytesIO
from typing import Dict

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def _col_to_index(col: str) -> int:
    result = 0
    for ch in col:
        result = result * 26 + (ord(ch.upper()) - ord("A") + 1)
    return result


def _clear_range_values(ws, start_row: int, end_row: int, start_col: str, end_col: str) -> None:
    min_col = _col_to_index(start_col)
    max_col = _col_to_index(end_col)
    for r in range(start_row, end_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(r, c).value = None


def _copy_cell_style(src_cell, dst_cell) -> None:
    dst_cell._style = copy(src_cell._style)
    dst_cell.number_format = src_cell.number_format
    dst_cell.font = copy(src_cell.font)
    dst_cell.fill = copy(src_cell.fill)
    dst_cell.border = copy(src_cell.border)
    dst_cell.alignment = copy(src_cell.alignment)
    dst_cell.protection = copy(src_cell.protection)


def _apply_row_style(ws, style_row: int, target_row: int, start_col: int, end_col: int) -> None:
    for col in range(start_col, end_col + 1):
        _copy_cell_style(ws.cell(style_row, col), ws.cell(target_row, col))
    ws.row_dimensions[target_row].height = ws.row_dimensions[style_row].height


def _remove_merged_ranges_on_rows(ws, min_row: int, max_row: int) -> None:
    merged_to_remove = []
    for merged in ws.merged_cells.ranges:
        if not (merged.max_row < min_row or merged.min_row > max_row):
            merged_to_remove.append(str(merged))
    for m in merged_to_remove:
        ws.unmerge_cells(m)


def _sorted_week_names(weekly_sheets: Dict[str, pd.DataFrame]) -> list[str]:
    return sorted(weekly_sheets.keys(), key=lambda x: int(str(x).replace("W", "")))


def _export_with_template(
    monthly_df: pd.DataFrame,
    weekly_sheets: Dict[str, pd.DataFrame],
    template_bytes: bytes,
    week_date_labels: dict[str, str] | None = None,
) -> bytes:
    wb = load_workbook(BytesIO(template_bytes))
    week_names = _sorted_week_names(weekly_sheets)
    agents = sorted(monthly_df["客服姓名"].astype(str).unique().tolist())

    # Ensure weekly sheets exist.
    base_week_sheet = None
    for name in wb.sheetnames:
        if name.startswith("W"):
            base_week_sheet = wb[name]
            break
    if base_week_sheet is None:
        raise ValueError("Template must contain at least one weekly sheet like 'W1'.")

    for week in week_names:
        if week not in wb.sheetnames:
            new_ws = wb.copy_worksheet(base_week_sheet)
            new_ws.title = week

    # Remove extra weekly sheets not present in current data (e.g., template has W2-W4).
    for sheet_name in list(wb.sheetnames):
        if sheet_name.startswith("W") and sheet_name not in week_names:
            del wb[sheet_name]

    # Fill weekly sheets in template layout: A:G.
    for week in week_names:
        ws = wb[week]
        weekly_style_row = 3
        weekly_max_col = 7
        if week_date_labels and week in week_date_labels:
            ws["A1"] = week_date_labels[week]
        else:
            ws["A1"] = week
        _clear_range_values(ws, start_row=3, end_row=max(ws.max_row, 500), start_col="A", end_col="G")

        week_df = weekly_sheets[week].copy()
        week_df["客服姓名"] = week_df["客服姓名"].astype(str)
        by_agent = {row["客服姓名"]: row for _, row in week_df.iterrows()}

        for i, agent in enumerate(agents, start=3):
            _apply_row_style(ws, weekly_style_row, i, 1, weekly_max_col)
            row = by_agent.get(agent)
            ws[f"A{i}"] = agent
            if row is None:
                ws[f"B{i}"] = 0.0
                ws[f"C{i}"] = ""
                ws[f"D{i}"] = 0.0
                ws[f"E{i}"] = 5
                ws[f"F{i}"] = -2400.0
                ws[f"G{i}"] = 0.0
            else:
                y = float(row["周实际工时 Y"])
                x = float(row["质检系数 X"])
                m = float(row["超标准时间 M"])
                ws[f"B{i}"] = y
                ws[f"C{i}"] = str(row["质检等级"])
                ws[f"D{i}"] = x * y
                ws[f"E{i}"] = 5
                ws[f"F{i}"] = m
                ws[f"G{i}"] = float(row["周奖励"])

        last_data_row = 2 + len(agents)
        if ws.max_row > last_data_row:
            ws.delete_rows(last_data_row + 1, ws.max_row - last_data_row)

    # Fill summary sheet.
    summary_name = "绩效表" if "绩效表" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[summary_name]
    # Capture style prototypes from template before clearing.
    proto = {
        "a2": copy(ws["A2"]),
        "a3": copy(ws["A3"]),
        "week_title": copy(ws["B2"]),
        "week_h1": copy(ws["B3"]),
        "week_h2": copy(ws["C3"]),
        "week_h3": copy(ws["D3"]),
        "rate_h": copy(ws["N3"]) if ws.max_column >= _col_to_index("N") else copy(ws["E3"]),
        "avg_h": copy(ws["R3"]) if ws.max_column >= _col_to_index("R") else copy(ws["F3"]),
        "total_h": copy(ws["S3"]) if ws.max_column >= _col_to_index("S") else copy(ws["G3"]),
        "name_row": copy(ws["A4"]),
        "week_d1": copy(ws["B4"]),
        "week_d2": copy(ws["C4"]),
        "week_d3": copy(ws["D4"]),
        "rate_d": copy(ws["N4"]) if ws.max_column >= _col_to_index("N") else copy(ws["E4"]),
        "avg_d": copy(ws["R4"]) if ws.max_column >= _col_to_index("R") else copy(ws["F4"]),
        "total_d": copy(ws["S4"]) if ws.max_column >= _col_to_index("S") else copy(ws["G4"]),
    }
    _remove_merged_ranges_on_rows(ws, min_row=2, max_row=3)
    _clear_range_values(ws, start_row=2, end_row=max(ws.max_row, 500), start_col="A", end_col="ZZ")

    # Dynamic summary header: each week uses 3 columns for 工时/出勤/超激励奖励
    ws["A2"] = "客服"
    ws["A3"] = ""
    _copy_cell_style(proto["a2"], ws["A2"])
    _copy_cell_style(proto["a3"], ws["A3"])
    ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)

    week_block_start_cols = {}
    for idx, week in enumerate(week_names):
        start_col = 2 + idx * 3
        end_col = start_col + 2
        week_block_start_cols[week] = start_col
        ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
        ws.cell(2, start_col).value = week
        ws.cell(3, start_col).value = "工时"
        ws.cell(3, start_col + 1).value = "出勤"
        ws.cell(3, start_col + 2).value = "超激励奖励"
        _copy_cell_style(proto["week_title"], ws.cell(2, start_col))
        _copy_cell_style(proto["week_h1"], ws.cell(3, start_col))
        _copy_cell_style(proto["week_h2"], ws.cell(3, start_col + 1))
        _copy_cell_style(proto["week_h3"], ws.cell(3, start_col + 2))

    rate_start_col = 2 + len(week_names) * 3
    week_rate_cols = {}
    for idx, week in enumerate(week_names):
        col = rate_start_col + idx
        week_rate_cols[week] = col
        ws.cell(3, col).value = f"绩效达成{week}"
        _copy_cell_style(proto["rate_h"], ws.cell(3, col))

    avg_col = rate_start_col + len(week_names)
    total_reward_col = avg_col + 1
    ws.cell(3, avg_col).value = "平均绩效系数"
    ws.cell(3, total_reward_col).value = "超激励绩效"
    _copy_cell_style(proto["avg_h"], ws.cell(3, avg_col))
    _copy_cell_style(proto["total_h"], ws.cell(3, total_reward_col))

    week_agent_map = {}
    for week in week_names:
        df = weekly_sheets[week].copy()
        df["客服姓名"] = df["客服姓名"].astype(str)
        week_agent_map[week] = {row["客服姓名"]: row for _, row in df.iterrows()}

    for i, agent in enumerate(agents, start=4):
        ws[f"A{i}"] = agent
        _copy_cell_style(proto["name_row"], ws[f"A{i}"])

        for week in week_names:
            row = week_agent_map[week].get(agent)
            y = 0.0 if row is None else float(row["周实际工时 Y"])
            x_factor = 0.0 if row is None else float(row["质检系数 X"])
            corrected = x_factor * y
            reward = 0.0 if row is None else float(row["周奖励"])

            base_idx = week_block_start_cols[week]
            # Summary "工时" uses QA-corrected hours.
            ws.cell(i, base_idx).value = corrected
            ws.cell(i, base_idx + 1).value = 5
            ws.cell(i, base_idx + 2).value = reward
            _copy_cell_style(proto["week_d1"], ws.cell(i, base_idx))
            _copy_cell_style(proto["week_d2"], ws.cell(i, base_idx + 1))
            _copy_cell_style(proto["week_d3"], ws.cell(i, base_idx + 2))

            y_col = get_column_letter(base_idx)
            attend_col = get_column_letter(base_idx + 1)
            ws.cell(i, week_rate_cols[week]).value = (
                f"=IF({attend_col}{i}*480=0,0,{y_col}{i}/({attend_col}{i}*480))"
            )
            _copy_cell_style(proto["rate_d"], ws.cell(i, week_rate_cols[week]))

        rate_cols = [get_column_letter(week_rate_cols[w]) for w in week_names]
        reward_cols = [get_column_letter(week_block_start_cols[w] + 2) for w in week_names]
        ws.cell(i, avg_col).value = (
            f"=IFERROR(AVERAGE({','.join([f'{c}{i}' for c in rate_cols])}),0)"
        )
        ws.cell(i, total_reward_col).value = f"=SUM({','.join([f'{c}{i}' for c in reward_cols])})"
        _copy_cell_style(proto["avg_d"], ws.cell(i, avg_col))
        _copy_cell_style(proto["total_d"], ws.cell(i, total_reward_col))

    last_summary_row = 3 + len(agents)
    if ws.max_row > last_summary_row:
        ws.delete_rows(last_summary_row + 1, ws.max_row - last_summary_row)

    # Basic width for appended columns.
    for c in range(1, total_reward_col + 1):
        letter = get_column_letter(c)
        if ws.column_dimensions[letter].width is None:
            ws.column_dimensions[letter].width = 14
    if ws.max_column > total_reward_col:
        ws.delete_cols(total_reward_col + 1, ws.max_column - total_reward_col)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def export_to_excel(
    monthly_df: pd.DataFrame,
    weekly_sheets: Dict[str, pd.DataFrame],
    template_bytes: bytes | None = None,
    week_date_labels: dict[str, str] | None = None,
) -> bytes:
    if template_bytes:
        return _export_with_template(
            monthly_df, weekly_sheets, template_bytes, week_date_labels=week_date_labels
        )

    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        monthly_df.to_excel(writer, index=False, sheet_name="绩效表（月汇总）")
        for week_name, week_df in weekly_sheets.items():
            week_df.to_excel(writer, index=False, sheet_name=week_name)

    output.seek(0)
    return output.getvalue()
